"""
Excel Reporter
Generates a professional, formatted Excel report using openpyxl.
Includes: Summary Dashboard + Detailed Findings sheets.
"""

import os
import platform
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint


# ── Colour palette ─────────────────────────────────────────────
C_COMPLIANT     = "FF00B050"   # Green
C_NON_COMPLIANT = "FFFF0000"   # Red
C_ERROR         = "FFFFC000"   # Amber
C_HEADER_BG     = "FF1F3864"   # Dark navy
C_HEADER_FG     = "FFFFFFFF"   # White
C_ALT_ROW       = "FFF2F2F2"   # Light grey
C_SECTION_BG    = "FFD6E4F0"   # Light blue
C_CRITICAL      = "FFFF0000"
C_HIGH          = "FFFF6600"
C_MEDIUM        = "FFFFC000"
C_LOW           = "FF00B050"

SEVERITY_COLOR  = {"Critical": C_CRITICAL, "High": C_HIGH, "Medium": C_MEDIUM, "Low": C_LOW}

THIN = Side(style="thin", color="FFD3D3D3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_BODY   = Font(name="Arial", size=10)
FONT_BOLD   = Font(name="Arial", size=10, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color=C_HEADER_FG)
FONT_TITLE  = Font(name="Arial", size=16, bold=True, color=C_HEADER_BG)
FONT_SUB    = Font(name="Arial", size=11, color="FF555555")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


class ExcelReporter:
    def __init__(self, results: List[Dict], system_type: str, depth: str, output_path: str):
        self.results = results
        self.system_type = system_type
        self.depth = depth
        self.output_path = output_path
        self.wb = Workbook()
        self.timestamp = datetime.now().strftime("%d %b %Y %H:%M")

    def generate(self):
        self._build_summary_sheet()
        self._build_findings_sheet()
        self._build_category_sheet()
        self.wb.save(self.output_path)

    # ──────────────────────────────────────────────────────────────
    # Sheet 1: Summary Dashboard
    # ──────────────────────────────────────────────────────────────
    def _build_summary_sheet(self):
        ws = self.wb.active
        ws.title = "Summary"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.row_dimensions[1].height = 10

        # Title block
        ws.merge_cells("B2:F3")
        title = ws["B2"]
        title.value = "CIS Benchmark Configuration Review"
        title.font = FONT_TITLE
        title.alignment = _left()

        ws.merge_cells("B4:F4")
        sub = ws["B4"]
        sub.value = (
            f"System: {self.system_type.replace('_', ' ').title()}  |  "
            f"Depth: {self.depth.capitalize()}  |  "
            f"Scan Date: {self.timestamp}  |  "
            f"Host: {platform.node()}"
        )
        sub.font = FONT_SUB
        sub.alignment = _left()
        ws.row_dimensions[4].height = 18

        # ── Stat boxes ───────────────────────────────────────────
        total        = len(self.results)
        compliant    = sum(1 for r in self.results if r["Status"] == "Compliant")
        non_compliant= sum(1 for r in self.results if r["Status"] == "Non-Compliant")
        errors       = sum(1 for r in self.results if r["Status"] == "Error")
        score        = round((compliant / total * 100) if total else 0, 1)

        stats = [
            ("Total Checks", total,         "FF1F3864"),
            ("Compliant",    compliant,      "FF00B050"),
            ("Non-Compliant",non_compliant,  "FFCC0000"),
            ("Errors",       errors,         "FFFFC000"),
        ]

        for col_idx, (label, value, color) in enumerate(stats, start=2):
            col = get_column_letter(col_idx)
            ws.row_dimensions[6].height = 22
            ws.row_dimensions[7].height = 38
            ws.row_dimensions[8].height = 22

            ws.merge_cells(f"{col}6:{col}6")
            lbl_cell = ws[f"{col}6"]
            lbl_cell.value = label
            lbl_cell.font = Font(name="Arial", size=9, bold=True, color="FF555555")
            lbl_cell.alignment = _center()
            lbl_cell.fill = _fill("FFF7F7F7")
            lbl_cell.border = BORDER

            ws.merge_cells(f"{col}7:{col}7")
            val_cell = ws[f"{col}7"]
            val_cell.value = value
            val_cell.font = Font(name="Arial", size=22, bold=True, color=color)
            val_cell.alignment = _center()
            val_cell.fill = _fill("FFFFFFFF")
            val_cell.border = BORDER

        # Compliance score
        ws.merge_cells("B9:F9")
        score_cell = ws["B9"]
        score_cell.value = f"Compliance Score: {score}%"
        score_cell.font = Font(name="Arial", size=13, bold=True,
                               color="FF00B050" if score >= 80 else ("FFFFC000" if score >= 60 else "FFCC0000"))
        score_cell.alignment = _center()
        score_cell.fill = _fill("FFF0FFF0" if score >= 80 else ("FFFFFFF0" if score >= 60 else "FFFFF0F0"))
        score_cell.border = BORDER
        ws.row_dimensions[9].height = 28

        # ── Severity breakdown ────────────────────────────────────
        ws.row_dimensions[11].height = 20
        ws.merge_cells("B11:F11")
        sev_hdr = ws["B11"]
        sev_hdr.value = "Non-Compliant Findings by Severity"
        sev_hdr.font = FONT_BOLD
        sev_hdr.alignment = _left()
        sev_hdr.fill = _fill(C_SECTION_BG)
        sev_hdr.border = BORDER

        severities = ["Critical", "High", "Medium", "Low"]
        nc_results = [r for r in self.results if r["Status"] == "Non-Compliant"]
        sev_counts = {s: sum(1 for r in nc_results if r.get("Severity") == s) for s in severities}

        for i, sev in enumerate(severities):
            row = 12 + i
            ws.row_dimensions[row].height = 20
            lbl = ws.cell(row=row, column=2, value=sev)
            lbl.font = Font(name="Arial", size=10, bold=True, color=SEVERITY_COLOR[sev])
            lbl.fill = _fill("FFFFFFFF")
            lbl.border = BORDER
            cnt = ws.cell(row=row, column=3, value=sev_counts[sev])
            cnt.font = FONT_BODY
            cnt.alignment = _center()
            cnt.border = BORDER

        # ── Category summary table ────────────────────────────────
        ws.row_dimensions[17].height = 20
        ws.merge_cells("B17:F17")
        cat_hdr = ws["B17"]
        cat_hdr.value = "Results by Category"
        cat_hdr.font = FONT_HEADER
        cat_hdr.fill = _fill(C_HEADER_BG)
        cat_hdr.alignment = _center()
        cat_hdr.border = BORDER

        col_headers = ["Category", "Total", "Compliant", "Non-Compliant", "Pass Rate"]
        for ci, h in enumerate(col_headers, start=2):
            cell = ws.cell(row=18, column=ci, value=h)
            cell.font = FONT_BOLD
            cell.fill = _fill(C_SECTION_BG)
            cell.alignment = _center()
            cell.border = BORDER
        ws.row_dimensions[18].height = 20

        categories = sorted(set(r["Category"] for r in self.results))
        for ri, cat in enumerate(categories):
            row = 19 + ri
            ws.row_dimensions[row].height = 18
            cat_results = [r for r in self.results if r["Category"] == cat]
            c_total = len(cat_results)
            c_comp  = sum(1 for r in cat_results if r["Status"] == "Compliant")
            c_nc    = sum(1 for r in cat_results if r["Status"] == "Non-Compliant")
            c_rate  = f"{round(c_comp/c_total*100) if c_total else 0}%"
            fill = _fill(C_ALT_ROW) if ri % 2 == 0 else _fill("FFFFFFFF")
            for ci, val in enumerate([cat, c_total, c_comp, c_nc, c_rate], start=2):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = FONT_BODY
                cell.fill = fill
                cell.alignment = _center() if ci > 2 else _left()
                cell.border = BORDER

    # ──────────────────────────────────────────────────────────────
    # Sheet 2: Detailed Findings
    # ──────────────────────────────────────────────────────────────
    def _build_findings_sheet(self):
        ws = self.wb.create_sheet("Findings")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        columns = [
            ("Control ID",      14),
            ("Control Name",    38),
            ("Category",        22),
            ("Severity",        12),
            ("Status",          16),
            ("Expected Value",  26),
            ("Observed Value",  26),
            ("Evidence",        42),
            ("Remediation",     52),
            ("References",      30),
        ]

        # Header row
        for col_idx, (header, width) in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = _fill(C_HEADER_BG)
            cell.alignment = _center(wrap=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 28

        # Data rows
        col_keys = [c[0] for c in columns]
        for row_idx, result in enumerate(self.results, start=2):
            status = result.get("Status", "")
            is_alt = (row_idx % 2 == 0)
            default_fill = _fill(C_ALT_ROW) if is_alt else _fill("FFFFFFFF")

            for col_idx, key in enumerate(col_keys, start=1):
                val = result.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = FONT_BODY
                cell.border = BORDER
                cell.alignment = _left(wrap=True)

                if key == "Status":
                    if status == "Compliant":
                        cell.fill = _fill(C_COMPLIANT)
                        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
                    elif status == "Non-Compliant":
                        cell.fill = _fill(C_NON_COMPLIANT)
                        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
                    elif status == "Error":
                        cell.fill = _fill(C_ERROR)
                        cell.font = Font(name="Arial", size=10, bold=True)
                    else:
                        cell.fill = default_fill
                    cell.alignment = _center()
                elif key == "Severity":
                    sev_color = SEVERITY_COLOR.get(val, "FF555555")
                    cell.font = Font(name="Arial", size=10, bold=True, color=sev_color)
                    cell.fill = default_fill
                    cell.alignment = _center()
                else:
                    cell.fill = default_fill

            ws.row_dimensions[row_idx].height = 30

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # ──────────────────────────────────────────────────────────────
    # Sheet 3: Category Breakdown
    # ──────────────────────────────────────────────────────────────
    def _build_category_sheet(self):
        ws = self.wb.create_sheet("By Category")
        ws.sheet_view.showGridLines = False

        categories = sorted(set(r["Category"] for r in self.results))
        headers = ["Category", "Total", "Compliant", "Non-Compliant", "Errors", "Pass Rate %"]

        col_widths = [30, 10, 12, 16, 10, 14]
        for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = FONT_HEADER
            cell.fill = _fill(C_HEADER_BG)
            cell.alignment = _center(wrap=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 28

        for ri, cat in enumerate(categories, start=2):
            cat_r = [r for r in self.results if r["Category"] == cat]
            total = len(cat_r)
            comp  = sum(1 for r in cat_r if r["Status"] == "Compliant")
            nc    = sum(1 for r in cat_r if r["Status"] == "Non-Compliant")
            err   = sum(1 for r in cat_r if r["Status"] == "Error")
            rate  = round(comp / total * 100, 1) if total else 0

            row_fill = _fill(C_ALT_ROW) if ri % 2 == 0 else _fill("FFFFFFFF")
            for ci, val in enumerate([cat, total, comp, nc, err, rate], start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = FONT_BODY
                cell.fill = row_fill
                cell.border = BORDER
                cell.alignment = _center() if ci > 1 else _left()
                if ci == 6:
                    cell.number_format = "0.0"
                    if rate >= 80:
                        cell.font = Font(name="Arial", size=10, bold=True, color="FF00B050")
                    elif rate >= 60:
                        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFC000")
                    else:
                        cell.font = Font(name="Arial", size=10, bold=True, color="FFCC0000")
            ws.row_dimensions[ri].height = 20
